# app1_golf_score_calculator/app.py
from datetime import datetime
import os
import re
from flask import Flask, session
from .routes import score_calc_bp, golf_bp
from .utils.excel_session import close_excel
from openpyxl import load_workbook
import xlrd
import mysql.connector
from utils.course import extract_course_name_and_date
from .excel_cache import ExcelCache
from utils.names import canonicalize_name

def mysql_connection():
    try:
        connection = mysql.connector.connect(
            host=os.getenv("DB_HOST"),
            user=os.getenv("DB_USER"),
            password=os.getenv("DB_PASSWORD"),
            database=os.getenv("DB_NAME")
        )
        print("✅ Database connection successful")
        return connection
    except mysql.connector.Error as err:
        print(f"❌ Database connection error: {err}")
        raise


def get_or_create_course_id(conn, course_name, tournament_date):
    cursor = conn.cursor()

    # Normalize date for comparison
    month_year = datetime.strptime(tournament_date, "%B %Y")
    start_date = month_year.replace(day=1).date()

    # ✅ Pull the most recent course with the matching name and date
    cursor.execute("""
        SELECT course_id FROM courses
        WHERE course_name = %s AND tournament_date = %s
        ORDER BY course_id DESC LIMIT 1
    """, (course_name, start_date))
    result = cursor.fetchone()

    if result:
        course_id = result[0]
        print(f"✅ Found existing course ID: {course_id}")
    else:
        cursor.execute("""
            INSERT INTO courses (course_name, tournament_date)
            VALUES (%s, %s)
        """, (course_name, start_date))
        conn.commit()
        course_id = cursor.lastrowid
        print(f"➕ Inserted new course ID: {course_id}")

    cursor.close()
    return course_id

def preload_players_from_excel():
    """Preload player names from an Excel file."""
    try:
        # Look for the Excel file in the root directory
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        target_file = None
        for file in os.listdir(base_dir):
            if file.endswith('.xlsx') and "scoring sheet" in file.lower():
                target_file = os.path.join(base_dir, file)
                print(f"🔍 Found sheet: {target_file}")
                break
        
        if not target_file:
            print("⚠️ No scoring sheet found")
            return []
            
        # Use openpyxl for xlsx files
        print(f"📊 Opening workbook: {target_file}")
        wb = load_workbook(target_file, read_only=True)
        sheet = wb.active
        print(f"📑 Using sheet: {sheet.title}")
        
        players = []
        
        # Helper function to filter invalid placeholder names
        def is_valid_player_name(first, last):
            if not first or not last:
                return False
                
            # Convert to strings and clean
            first = str(first).strip().lower()
            last = str(last).strip().lower()
            
            # Check for empty values
            if not first or not last:
                return False
                
            # Check for headers
            if first == "first" or first == "first name" or last == "last" or last == "last name":
                return False
                
            # Check for placeholders (e.g., "Xxxxxxxx")
            if re.match(r'^x+$', first, re.IGNORECASE) or re.match(r'^x+$', last, re.IGNORECASE):
                return False
            
            # Check for numeric values
            if first.isdigit() or last.isdigit():
                return False
            
            # Check for common placeholder values
            placeholders = ['none', 'null', 'na', 'n/a', 'blank', 'unknown', 'test']
            if first in placeholders or last in placeholders:
                return False
                
            return True
        
        # Process players - read both first and last name columns if available
        for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0]:  # Check if first name cell is not empty
                first_name = str(row[0]).strip()
                last_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                
                # Filter out placeholders and headers
                if is_valid_player_name(first_name, last_name):
                    full_name = f"{first_name} {last_name}".strip()
                    if full_name and len(full_name) > 2:  # Basic validation
                        players.append(full_name)
                        print(f"👤 Found player: {full_name}")
                else:
                    print(f"⚠️ Skipping invalid player name: {first_name} {last_name}")
        
        if not players:
            print("⚠️ No valid players found in the Excel file")
            return []
            
        # Connect to the database
        print("🔌 Connecting to database")
        conn = mysql_connection()
        cursor = conn.cursor()
        
        # Check table structure
        cursor.execute("DESCRIBE players")
        columns = [row[0].lower() for row in cursor.fetchall()]
        
        # Determine the correct column name
        name_column = None
        potential_columns = ['name', 'player_name', 'full_name', 'first_name']
        for col in potential_columns:
            if col in columns:
                name_column = col
                break
        
        if not name_column:
            # If none of our guesses match, try to create one
            print("⚠️ No suitable name column found in players table")
            try:
                cursor.execute("ALTER TABLE players ADD COLUMN name VARCHAR(255)")
                name_column = 'name'
                print("✅ Added 'name' column to players table")
            except:
                print("❌ Failed to add column to players table")
                return []
        
        # Get current course_id
        cursor.execute("SELECT COALESCE(MAX(course_id), 1) FROM courses")
        course_id = cursor.fetchone()[0]

        # Insert or update players (don't truncate!)
        inserted = 0
        updated = 0
        for name in players:
            # Split name into first and last
            parts = name.split(' ', 1)
            first_name = parts[0]
            last_name = parts[1] if len(parts) > 1 else ''

            # Check if player already exists for this course
            cursor.execute("""
                SELECT id FROM players
                WHERE first_name = %s AND last_name = %s AND course_id = %s
            """, (first_name, last_name, course_id))

            if cursor.fetchone():
                updated += 1  # Already exists, skip
            else:
                # Insert new player
                cursor.execute("""
                    INSERT INTO players (first_name, last_name, course_id)
                    VALUES (%s, %s, %s)
                """, (first_name, last_name, course_id))
                inserted += 1
        
        conn.commit()
        cursor.close()
        conn.close()

        print(f"✅ Preloaded players from Excel: {inserted} new, {updated} existing")
        return players
    except Exception as e:
        print(f"❌ Error preloading players: {e}")
        import traceback
        traceback.print_exc()
        return []

def create_app():
    app = Flask(__name__)
    app.secret_key = os.getenv("SECRET_KEY", os.urandom(24).hex())
    
    # Preload players right at app initialization
    player_list = preload_players_from_excel()
    print(f"🚀 App started with {len(player_list) if player_list else 0} preloaded players")

    @app.before_request
    def set_course():
        # only do this once per session
        if 'course_name' not in session:
            # find the path to the open scoring sheet
            excel_path = ExcelCache.get_excel_path()
            # parse it in one shot
            name, date = extract_course_name_and_date(excel_path)
            session['course_name'] = name
            session['course_date'] = date

    app.register_blueprint(score_calc_bp)
    app.register_blueprint(golf_bp)  # Register this blueprint here instead of after creating the app

    return app

app = create_app()